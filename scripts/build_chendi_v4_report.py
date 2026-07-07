#!/usr/bin/env python3
"""Build spreadsheet per Chendi's spec (2026-07-07):

Each row = one (model, config).
Columns:
  - model
  - full knob values (7 knobs, unique for this config)
  - batch=1 metrics: tokens/s, MFU_amort, MFU_simple, MBU, HBM_used_peak_GB, decode_step_ms
  - batch=32 metrics: same
  - Optional extras: batch=64, batch=128 columns for trend

Sheets: one per regime family
  - Sheet 1: "R_decode_c1_out2k"  (batch=1, long output)
  - Sheet 2: "R_conc_ref"          (batch=32, short output)
  - Sheet 3: "R_decode_c32_out1k"  (batch=32, longer output)
  - Sheet 4: "R_decode_c64_out512" (batch=64)
  - Sheet 5: "R_decode_c128_out256" (batch=128)
  - Sheet 6: "combined" (all metrics side-by-side, chendi's requested view)

Includes new LFM2.5-8B-A1B rows added 2026-07-07.
"""
from __future__ import annotations
import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

REPO = Path(__file__).resolve().parent.parent
V4 = REPO / "results/2026-07-07_v4_decode_sweep"

MODELS = [
    "lfm2.5-8b-a1b",
    "qwen3-30b-a3b-bf16",
    "qwen3-30b-a3b-fp8",
    "qwen3-0.6b",
]

CONFIGS = ["cookbook_baseline", "v3_best_chunk8k", "big_batch_cap128"]

# Regime → (label, prompt_words, out_tok, concurrency)
REGIMES = {
    "R_decode_c1_out2k":    ("c1_out2k",    100, 2048, 1),
    "R_conc_ref":           ("c32_out256",  200, 256,  32),
    "R_decode_c32_out1k":   ("c32_out1k",   200, 1024, 32),
    "R_decode_c64_out512":  ("c64_out512",  200, 512,  64),
    "R_decode_c128_out256": ("c128_out256", 200, 256,  128),
}


def load_all():
    out = {}
    for m in MODELS:
        for c in CONFIGS:
            sp = V4 / m / c / "summary.json"
            hp = V4 / m / c / "hw_stats.json"
            if not sp.exists() or not hp.exists():
                continue
            out[(m, c)] = {
                "summary": json.load(open(sp)),
                "hw": json.load(open(hp)),
            }
    return out


def get_knobs(spec_dict):
    """Extract the 7 tunable knobs from the resolved spec."""
    sc = spec_dict.get("spec_resolved", {}).get("server_config", {})
    return {
        "moe_runner_backend": sc.get("moe-runner-backend", "auto"),
        "attention_backend":  sc.get("attention-backend", "auto"),
        "cuda_graph":         "off" if sc.get("disable-cuda-graph") else "on",
        "max_running_reqs":   sc.get("max-running-requests"),
        "chunked_prefill":    sc.get("chunked-prefill-size"),
        "schedule_policy":    sc.get("schedule-policy"),
        "mem_fraction_static": sc.get("mem-fraction-static"),
    }


def regime_row(model, config, data):
    """Extract per-regime metrics for one (model, config)."""
    knobs = get_knobs(data["summary"])
    hw_stats = data["hw"].get("gpu_stats", {})
    sglang_timings = data["hw"].get("sglang_timings", {})
    hbm_used_peak_gb = hw_stats.get("memory_used_mib", {}).get("max", 0) / 1024
    hbm_bw_peak_pct = hw_stats.get("memory_util_pct", {}).get("max", 0)
    power_peak_w = hw_stats.get("power_draw_w", {}).get("max", 0)

    row = {
        "model": model,
        "config": config,
        **knobs,
        "hbm_used_peak_GB": round(hbm_used_peak_gb, 1),
        "hbm_bw_peak_pct": hbm_bw_peak_pct,
        "power_peak_W": round(power_peak_w),
    }

    # Per-regime metrics
    for reg_id, (short, pw, ot, conc) in REGIMES.items():
        reg = data["summary"].get("regimes", {}).get(reg_id, {})
        prefix = short
        if reg:
            mfu = reg.get("mfu", {})
            row[f"{prefix}__tokens_per_s"] = round(reg["tokens_per_s"]["mean"], 1)
            row[f"{prefix}__req_per_s"] = round(reg["req_per_s"]["mean"], 3)
            row[f"{prefix}__MFU_amort_pct"] = round(mfu.get("mfu_pct_amortized", 0), 2)
            row[f"{prefix}__MFU_simple_pct"] = round(mfu.get("mfu_pct_simple", 0), 3)
            row[f"{prefix}__MBU_pct"] = round(mfu.get("mbu_pct", 0), 2)
            # TPOT = time per output token per request (ms)
            # = 1 / (tokens_per_s_per_request) = concurrency / tokens_per_s * 1000
            tps = reg["tokens_per_s"]["mean"]
            tpot_ms = 1000.0 * conc / tps if tps > 0 else 0
            row[f"{prefix}__TPOT_ms"] = round(tpot_ms, 3)
            # decode step time at this batch (from sglang log)
            step_info = sglang_timings.get(f"batch_{conc}", {})
            row[f"{prefix}__decode_step_ms"] = step_info.get("median_step_ms", "")
        else:
            for suffix in ["tokens_per_s", "req_per_s", "MFU_amort_pct",
                           "MFU_simple_pct", "MBU_pct", "TPOT_ms", "decode_step_ms"]:
                row[f"{prefix}__{suffix}"] = ""
    return row


def build_workbook(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # README sheet
    ws = wb.create_sheet("README", 0)
    readme = [
        ("Cross-model decode-stress sweep — per Chendi's request 2026-07-07", "bold"),
        ("", None),
        ("Hardware: 1× NVIDIA H200 (141 GB HBM3e, 4.8 TB/s peak, 989 TFLOPS bf16 / 1979 TFLOPS fp8)", None),
        ("Software: sglang v0.5.9 + local autotune patch, conda env sglang-dev", None),
        ("", None),
        ("Rows: one per (model × config) combination.", None),
        ("Columns:", "bold"),
        ("  meta:    model, config, 7 tunable knobs (unique per config)", None),
        ("  system:  hbm_used_peak_GB (nvidia-smi memory.used max),", None),
        ("           hbm_bw_peak_pct (nvidia-smi utilization.memory max = dram_util),", None),
        ("           power_peak_W (nvidia-smi power.draw max)", None),
        ("  per-regime (5 regimes, prefix like 'c32_out256'):", None),
        ("    tokens_per_s, req_per_s = MEASURED bench output", None),
        ("    MFU_amort_pct    = derived FLOPs (prefill + decode) / peak_flops", None),
        ("    MFU_simple_pct   = derived FLOPs (decode only) / peak_flops", None),
        ("    MBU_pct          = derived weight_bytes × fwd_passes/s / peak_HBM_BW", None),
        ("    TPOT_ms          = time per output token per request = concurrency / tokens_per_s × 1000", None),
        ("    decode_step_ms   = REAL median decode step time from sglang gen_throughput log", None),
        ("", None),
        ("The 5 regimes:", "bold"),
        ("  c1_out2k     = 100 word prompt, 2048 output tokens, concurrency=1", None),
        ("  c32_out256   = 200 word prompt, 256 output tokens, concurrency=32  (reference from v3)", None),
        ("  c32_out1k    = 200 word prompt, 1024 output tokens, concurrency=32", None),
        ("  c64_out512   = 200 word prompt, 512 output tokens, concurrency=64", None),
        ("  c128_out256  = 200 word prompt, 256 output tokens, concurrency=128", None),
        ("", None),
        ("Sanity checks worth doing:", "bold"),
        ("  1. decode_step_ms should be roughly constant across configs for same (model, batch)", None),
        ("     — it reflects HBM read time which config can't change", None),
        ("  2. hbm_bw_peak_pct should be similar within a model across configs (~69% bf16, ~54% fp8)", None),
        ("     — it's the memory ceiling", None),
        ("  3. tokens_per_s should scale roughly linearly with concurrency until KV cache OOMs", None),
        ("     — degradation at larger batch = KV pressure or scheduler overhead", None),
        ("  4. MFU_simple always < MFU_amort; long-prefill regimes see biggest gap (10-100x)", None),
        ("     — this is why MFU_simple is misleading for long prompts", None),
    ]
    for i, (txt, style) in enumerate(readme, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if style == "bold":
            c.font = Font(bold=True, size=11)
    ws.column_dimensions["A"].width = 110

    # Main data sheet: all configs, all regimes
    ws = wb.create_sheet("all_configs_all_regimes")
    if not rows:
        return
    headers = list(rows[0].keys())
    hdr_fill = PatternFill("solid", fgColor="4A4A4A")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            ws.cell(row=ri, column=ci, value=v)
    ws.freeze_panes = "C2"  # freeze model + config columns

    # Format numeric columns
    for ci, h in enumerate(headers, 1):
        col = get_column_letter(ci)
        if "pct" in h.lower():
            for ri in range(2, len(rows) + 2):
                cell = ws[f"{col}{ri}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'
        elif h.endswith("_GB"):
            for ri in range(2, len(rows) + 2):
                cell = ws[f"{col}{ri}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0" GB"'
        elif h.endswith("_ms"):
            for ri in range(2, len(rows) + 2):
                cell = ws[f"{col}{ri}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00" ms"'
        elif h.endswith("_W"):
            for ri in range(2, len(rows) + 2):
                cell = ws[f"{col}{ri}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0" W"'
        # Width heuristic
        max_len = len(h)
        for ri in range(2, min(len(rows)+2, 30)):
            v = ws.cell(row=ri, column=ci).value
            if v is None: continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(max_len + 2, 20)
    ws.row_dimensions[1].height = 40

    # Color-scale tokens/s columns
    for ci, h in enumerate(headers, 1):
        if h.endswith("__tokens_per_s"):
            col = get_column_letter(ci)
            rng = f"{col}2:{col}{len(rows)+1}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            )
            ws.conditional_formatting.add(rng, rule)

    # Per-regime split sheets (chendi's original "one sheet per regime" style)
    for reg_id, (short, pw, ot, conc) in REGIMES.items():
        ws = wb.create_sheet(f"{short}")
        # Cols: meta + knobs + this regime's metrics
        cols = ["model", "config", "moe_runner_backend", "attention_backend",
                "cuda_graph", "max_running_reqs", "chunked_prefill",
                "schedule_policy", "mem_fraction_static",
                "hbm_used_peak_GB", "hbm_bw_peak_pct", "power_peak_W",
                f"{short}__tokens_per_s", f"{short}__req_per_s",
                f"{short}__TPOT_ms", f"{short}__decode_step_ms",
                f"{short}__MFU_amort_pct", f"{short}__MFU_simple_pct",
                f"{short}__MBU_pct"]
        # Regime info at top
        c = ws.cell(row=1, column=1,
                    value=f"Regime: {reg_id} — prompt={pw}w, out={ot} tok, concurrency={conc}")
        c.font = Font(bold=True, size=12)
        # Data starts row 3
        for ci, h in enumerate(cols, 1):
            hc = ws.cell(row=3, column=ci, value=h)
            hc.fill = hdr_fill; hc.font = hdr_font
            hc.alignment = Alignment(horizontal="center", wrap_text=True)
        for ri, row in enumerate(rows, 4):
            for ci, h in enumerate(cols, 1):
                v = row.get(h, "")
                ws.cell(row=ri, column=ci, value=v)
        ws.freeze_panes = "C4"

        # Format numeric cells
        for ci, h in enumerate(cols, 1):
            col = get_column_letter(ci)
            fmt = None
            if "pct" in h.lower(): fmt = '0.00"%"'
            elif h.endswith("_GB"): fmt = '0.0" GB"'
            elif h.endswith("_ms"): fmt = '0.00" ms"'
            elif h.endswith("_W"): fmt = '0" W"'
            elif h.endswith("_per_s"): fmt = '0.0'
            if fmt:
                for ri in range(4, len(rows) + 4):
                    cell = ws[f"{col}{ri}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = fmt
            # width
            max_len = len(h)
            for ri in range(4, len(rows) + 4):
                v = ws.cell(row=ri, column=ci).value
                if v is None: continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col].width = min(max_len + 2, 20)
        ws.row_dimensions[3].height = 40

    wb.save(out_path)


def main() -> int:
    data = load_all()
    rows = []
    for m in MODELS:
        for c in CONFIGS:
            if (m, c) not in data: continue
            rows.append(regime_row(m, c, data[(m, c)]))

    # Also write CSV
    csv_path = REPO / "results/consolidated_v4_by_model_config.csv"
    if rows:
        with csv_path.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    print(f"Wrote {len(rows)} rows → {csv_path}")

    xlsx_path = REPO / "results/chendi_v4_report.xlsx"
    build_workbook(rows, xlsx_path)
    print(f"Wrote {xlsx_path}")

    # Print top few for sanity
    print("\nSample rows:")
    for r in rows[:3]:
        print(f"  {r['model']:<22} {r['config']:<20} cap={r['max_running_reqs']}  chunk={r['chunked_prefill']}  mem={r['mem_fraction_static']}")
    print(f"\nTotal: {len(rows)} rows across {len(MODELS)} models × {len(CONFIGS)} configs")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
