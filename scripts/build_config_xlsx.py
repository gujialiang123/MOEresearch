#!/usr/bin/env python3
"""Convert consolidated_config_spreadsheet.csv → xlsx with formatting.

Creates a workbook with two sheets:
  1. "sorted by R_conc speedup"  ← primary tuning target
  2. "sorted by geomean speedup" ← overall (long-context matters too)

Formatting:
  - Frozen header row + frozen first 4 meta columns
  - Speedup cells color-scaled (red < 1.0, green > 1.0, brightness ∝ magnitude)
  - Speedup columns display as e.g. "1.36×"
  - Auto-fit column widths

Usage:
    python scripts/build_config_xlsx.py \\
        --csv results/consolidated_config_spreadsheet.csv \\
        --out results/consolidated_config_spreadsheet.xlsx
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent


def load_rows(csv_path: Path) -> tuple[list[str], list[dict]]:
    with csv_path.open() as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        return list(reader.fieldnames), rows


def try_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def add_sheet(wb: Workbook, sheet_name: str, headers: list[str],
              rows: list[dict], sort_by: str) -> None:
    # Sort
    def key(r):
        v = try_float(r.get(sort_by))
        return (0 if v is None else 1, v if v is not None else 0)
    rows_sorted = sorted(rows, key=key, reverse=True)

    ws = wb.create_sheet(sheet_name)

    # Header row
    header_fill = PatternFill("solid", fgColor="4A4A4A")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    speedup_cols = [i for i, h in enumerate(headers, 1)
                    if "speedup" in h.lower()]
    for ri, row in enumerate(rows_sorted, 2):
        for ci, h in enumerate(headers, 1):
            val = row.get(h, "")
            f = try_float(val)
            if f is not None and val != "":
                ws.cell(row=ri, column=ci, value=f)
            else:
                ws.cell(row=ri, column=ci, value=val)

    # Freeze pane: header row + first 4 meta cols
    ws.freeze_panes = "E2"

    # Format speedup columns as "1.36×"
    for col_i in speedup_cols:
        col_letter = get_column_letter(col_i)
        for row_i in range(2, len(rows_sorted) + 2):
            cell = ws[f"{col_letter}{row_i}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00"×"'

    # Color-scale speedup cells: 0.5 = red, 1.0 = white, 1.5 = green, 2.5 = deep green
    for col_i in speedup_cols:
        col_letter = get_column_letter(col_i)
        rng = f"{col_letter}2:{col_letter}{len(rows_sorted)+1}"
        rule = ColorScaleRule(
            start_type="num", start_value=0.3, start_color="F8696B",  # red
            mid_type="num",   mid_value=1.0,  mid_color="FFFFFF",     # white
            end_type="num",   end_value=2.5,  end_color="63BE7B",     # green
        )
        ws.conditional_formatting.add(rng, rule)

    # MFU and MBU columns: format as "1.23%"
    for ci, h in enumerate(headers, 1):
        if ("MFU" in h or "MBU" in h) and "pct" in h:
            col_letter = get_column_letter(ci)
            for row_i in range(2, len(rows_sorted) + 2):
                cell = ws[f"{col_letter}{row_i}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'

    # Auto-fit width (heuristic)
    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        # header width + slight padding
        max_len = len(h)
        for row_i in range(2, min(len(rows_sorted) + 2, 100)):
            v = ws.cell(row=row_i, column=ci).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        # cap widths: meta cols wide, numeric cols narrow
        if "notes" in h.lower():
            width = min(max_len + 2, 45)
        elif "run_id" in h.lower():
            width = min(max_len + 2, 32)
        elif "spec_hash" in h.lower():
            width = 12
        else:
            width = min(max_len + 2, 18)
        ws.column_dimensions[col_letter].width = max(8, width)

    # Row height for header
    ws.row_dimensions[1].height = 30

    print(f"  Sheet '{sheet_name}': {len(rows_sorted)} data rows")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--csv", default="results/consolidated_config_spreadsheet.csv")
    ap.add_argument("--out", default="results/consolidated_config_spreadsheet.xlsx")
    args = ap.parse_args()

    csv_path = REPO / args.csv if not Path(args.csv).is_absolute() else Path(args.csv)
    out_path = REPO / args.out if not Path(args.out).is_absolute() else Path(args.out)

    # Load CSV
    with csv_path.open() as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames)
        rows = list(reader)
    print(f"Loaded {len(rows)} rows × {len(headers)} cols from {csv_path}")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # README sheet
    ws_readme = wb.create_sheet("README", 0)
    readme_lines = [
        ("LFM2.5-8B-A1B on 1× H200 — consolidated sglang config sweep", None),
        ("", None),
        ("What this file is:", "bold"),
        ("  One row = one sglang server config that was benchmarked.", None),
        ("  Data from three experiment rounds:", None),
        ("    v2 (2026-06-30): 25-trial Optuna + 7 diagnostic archived trials + 3-run baseline reruns", None),
        ("    v3 (2026-07-02): 30-trial Optuna (4 warm-start + 26 TPE) + baselines", None),
        ("  All configs benchmarked with same regime harness (num_runs=3 per regime).", None),
        ("", None),
        ("Speedup baseline (denominator):", "bold"),
        ("  v3-baseline-cookbook single lifetime × 3 runs.", None),
        ("  This is the sglang cookbook recommended default:", None),
        ("    moe=auto (resolves to triton on H200 bf16), attn=fa3 (auto), cap=32,", None),
        ("    chunk=-1, sched=lpm, mem=0.85, cg-on + reasoning-parser qwen3 + tool-call-parser lfm2", None),
        ("  Cross-lifetime stddev of this baseline (from v2 3-run rerun) = 0.5% on R_conc.", None),
        ("", None),
        ("Sheets:", "bold"),
        ("  1. by R_conc speedup    → sorted by target regime (R_concurrent_decode) speedup", None),
        ("                             This is what Optuna was optimizing.", None),
        ("  2. by geomean speedup   → sorted by geometric mean speedup across all 8 regimes", None),
        ("                             This reflects the 'well-rounded' config, including long-context.", None),
        ("", None),
        ("Column groups:", "bold"),
        ("  meta        : source (v2/v3), experiment, run_id, notes, spec_hash", None),
        ("  knobs       : 7 sglang server flags that were tuned", None),
        ("  aggregate   : speedup_R_conc, speedup_geomean_all_regimes, n_regimes_measured", None),
        ("  per-regime  : 6 columns per regime: speedup, req/s, tokens/s,", None),
        ("                MFU_simple%, MFU_amortized%, MBU%", None),
        ("                Which utilization number to look at depends on the regime:", None),
        ("                  * short decode: look at MBU (memory-bound)", None),
        ("                  * long prefill: look at MFU_amortized (compute-bound)", None),
        ("                  * MFU_simple = decode-only matmul FLOPs (LOW for long-prefill regimes;", None),
        ("                    do not use for those — it undercounts by 100-1000×)", None),
        ("                Regimes ordered short → long: R_short_decode, R_medium, R_conc, R_long_prefill,", None),
        ("                R_prompt_8k, R_prompt_16k, R_prompt_32k, R_prompt_50k", None),
        ("", None),
        ("Color coding:", "bold"),
        ("  Speedup cells: red < 1.0 (worse than baseline), white = baseline, green > 1.0 (better)", None),
        ("  Empty cells: regime not measured for this config (v2 has 4/8 regime coverage)", None),
        ("", None),
        ("Key findings from these runs (see docs/2026-07-02/lfm2.5_v3_mfu_longctx.md for full report):", "bold"),
        ("  1. R_conc winning config: v3-trial-0019 = 1.00× baseline (i.e. cookbook already optimal on R_conc)", None),
        ("  2. Overall (geomean) winner: v3-trial-0029 = 1.37× baseline, biggest gains on long-context:", None),
        ("       R_prompt_50k: 2.39× baseline (2.4× tokens/s on 65k-token prompts)", None),
        ("       R_prompt_16k: 1.70×", None),
        ("       R_prompt_8k:  1.46×", None),
        ("  3. Key knob: chunked-prefill-size. Cookbook default '-1' (no chunking) blocks long prompts;", None),
        ("     switching to 2048 or 8192 gives 40-140% speedup on long regimes with zero R_conc regression.", None),
        ("  4. v2 Optuna best (v2-main-trial-0017) hit R_conc=0.94× (a 6% regression from cookbook)", None),
        ("     because pure TPE missed triton MoE. v3 warm-start fixed this: v3-trial-0001 = 1.00×.", None),
        ("", None),
        ("Contact:", "bold"),
        ("  gujialiang123@github.com", None),
        ("  Repo: EndtoEnd-auto-optimization", None),
        ("  Full report: docs/2026-07-02/lfm2.5_v3_mfu_longctx.md", None),
    ]
    for ri, (text, style) in enumerate(readme_lines, 1):
        c = ws_readme.cell(row=ri, column=1, value=text)
        if style == "bold":
            c.font = Font(bold=True, size=11)
    ws_readme.column_dimensions["A"].width = 120

    # Sheet 1: sorted by R_conc speedup
    add_sheet(wb, "by R_conc speedup", headers, rows, "speedup_R_conc")
    # Sheet 2: sorted by geomean speedup
    add_sheet(wb, "by geomean speedup", headers, rows, "speedup_geomean_all_regimes")

    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
