#!/usr/bin/env python3
"""Build unified NCU kernel report combining:
  - June 9 data: Qwen3-30B-A3B (bf16), 1 config (cookbook), 4 regimes
    Full --set full metrics, real sglang kernels.
    Source: results/2026-06-09_sglang_triton_sweep/ncu/*/ncu_summary.json
  - v6 (July 8): LFM2.5-8B-A1B, cookbook, 3 regimes
    Reduced sections (SpeedOfLight + Occupancy + LaunchStats) but same
    methodology (sglang.bench_one_batch + NCU).
    Source: results/2026-07-08_v6_ncu/lfm2.5-8b-a1b/cookbook_baseline/*/ncu_raw.csv

Output:
  results/consolidated_v6_sglang_ncu.csv (rows: one per (model, config, regime, kernel))
  results/v6_sglang_ncu_report.xlsx
"""
from __future__ import annotations
import csv, json, re
from collections import defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent


def short_name(kname: str) -> str:
    kl = kname.lower()
    if 'fused_moe_kernel' in kl: return 'fused_moe_kernel'
    if 'moe_sum_reduce' in kl:  return 'moe_sum_reduce'
    if 'flash::flashattnfwdcombine' in kl or 'FlashAttnFwdCombine' in kname: return 'flash_attn_fwd_combine'
    if 'cutlass' in kl and 'flash::enable' in kl: return 'flash_attn_main'
    if 'flash_fwd_splitkv_combine' in kl: return 'flash_fwd_splitkv_combine'
    if 'flash_fwd_splitkv' in kl:   return 'flash_fwd_splitkv'
    if 'flash_fwd' in kl:           return 'flash_fwd'
    if 'norm::fusedaddrmsnorm' in kl or 'FusedAddRMSNorm' in kname: return 'FusedAddRMSNorm'
    if 'norm::rmsnorm' in kl or 'RMSNormKernel' in kname:           return 'RMSNorm'
    if 'act_and_mul' in kl:  return 'act_and_mul'
    if 'topkgatingsoftmax' in kl: return 'topkGatingSoftmax'
    if 'batchqkapplyrotary' in kl or 'ApplyRotary' in kname: return 'rope'
    if 'cublaslt' in kl: return 'cublaslt_' + (kname.split('cublasLt::')[-1].split('<')[0] if 'cublasLt::' in kname else 'other')
    if 'nvjet_tst' in kl:
        m = re.search(r'nvjet_tst_(\w+?)_(\w+?)_', kname)
        return f'nvjet_gemm_{m.group(1)}_{m.group(2)}' if m else 'nvjet_gemm'
    if 'triton_poi' in kl: return 'triton_' + kl.split('triton_')[-1][:25]
    if 'triton_per' in kl: return 'triton_' + kl.split('triton_')[-1][:25]
    if 'conv1d_update' in kl: return 'causal_conv1d_update'
    if 'moe_align_block' in kl: return 'moe_align_block'
    if 'fused_qknorm' in kl: return 'fused_qknorm'
    if 'flashinfer' in kl and 'gemm' in kl: return 'flashinfer_gemm'
    if 'flashinfer' in kl: return 'flashinfer_other'
    if 'elementwise' in kl: return 'elementwise'
    if 'reduce' in kl: return 'reduce'
    if 'copy' in kl: return 'memcpy'
    if 'index' in kl: return 'index'
    if 'devicescan' in kl: return 'device_scan'
    return kname[:35]


def load_june_qwen3_data() -> list[dict]:
    """Parse the 4 June 9 ncu_summary.json files."""
    rows = []
    JUNE = REPO / 'results/2026-06-09_sglang_triton_sweep/ncu'
    REGIME_INFO = {
        'R_short_decode':      (1, 100, 256, 'decode'),
        'R_medium_balanced':   (8, 800, 256, 'decode'),
        'R_long_prefill':      (4, 4000, 32, 'prefill'),
        'R_concurrent_decode': (32, 200, 256, 'decode'),
    }
    for reg, (batch, in_w, out_t, stage) in REGIME_INFO.items():
        p = JUNE / reg / 'ncu_summary.json'
        if not p.exists():
            continue
        d = json.load(open(p))
        for k in d.get('kernels', []):
            m = k.get('metrics', {})
            rows.append({
                'model': 'qwen3-30b-a3b-bf16',
                'config': 'cookbook_baseline',
                'regime': reg,
                'batch_size': batch,
                'prompt_words': in_w,
                'output_tokens': out_t,
                'stage': stage,
                'kernel_short': short_name(k['kernel']),
                'kernel_full': k['kernel'][:80],
                'ncu_data_source': 'june-9-full-set',
                'sm_throughput_pct': round(m.get('sm_throughput_pct', 0), 2),
                'dram_throughput_pct': round(m.get('dram_throughput_pct', 0), 2),
                'tensor_pipe_active_pct': round(m.get('tensor_pipe_active_pct', 0), 2),
                'warps_active_pct': round(m.get('warps_active_pct', 0), 2),
                'l1_hit_pct': round(m.get('l1_hit_pct', 0), 2),
                'l2_hit_pct': round(m.get('l2_hit_pct', 0), 2),
                'stall_long_scoreboard_avg': round(m.get('stall_long_scoreboard_avg', 0), 2),
                'stall_math_pipe_throttle_avg': round(m.get('stall_math_pipe_throttle_avg', 0), 2),
                'verdict': k.get('verdict', ''),
                'headroom_estimate_pct': k.get('headroom_estimate_pct', ''),
            })
    return rows


def parse_ncu_csv_v6(csv_path: Path) -> dict[str, dict]:
    """Parse v6 ncu_raw.csv, pivot to one dict per kernel launch."""
    if not csv_path.exists():
        return {}
    text = csv_path.read_text()
    lines = text.splitlines()
    header_idx = None
    for i, l in enumerate(lines):
        if l.startswith('"ID"'):
            header_idx = i
            break
    if header_idx is None:
        return {}
    csv_text = '\n'.join(lines[header_idx:])
    reader = csv.DictReader(csv_text.splitlines())
    launches = defaultdict(lambda: {"metrics": {}})
    METRIC_MAP = {
        'DRAM Throughput': 'dram_throughput_pct',
        'Compute (SM) Throughput': 'sm_throughput_pct',
        'L1/TEX Cache Throughput': 'l1_throughput_pct',
        'L2 Cache Throughput': 'l2_throughput_pct',
        'Achieved Occupancy': 'warps_active_pct',
        'Duration': 'duration_ns',
    }
    for row in reader:
        kid = row.get('ID', '?')
        k = row.get('Kernel Name', '?')
        m = row.get('Metric Name', '')
        v = row.get('Metric Value', '')
        launch_key = (kid, k)
        launches[launch_key]['kernel'] = k
        if m in METRIC_MAP:
            try:
                launches[launch_key]['metrics'][METRIC_MAP[m]] = float(v)
            except (ValueError, TypeError):
                pass
    return dict(launches)


def load_v6_lfm_data() -> list[dict]:
    """Parse v6 LFM2.5 ncu_raw.csv files (if they exist yet)."""
    rows = []
    V6 = REPO / 'results/2026-07-08_v6_ncu/lfm2.5-8b-a1b/cookbook_baseline'
    REGIME_INFO = {
        'R_decode_c1_out2k':    (1,   100, 2048, 'decode'),
        'R_conc_ref':           (32,  200, 256,  'decode'),
        'R_decode_c128_out256': (128, 200, 256,  'decode'),
    }
    for reg, (batch, in_w, out_t, stage) in REGIME_INFO.items():
        csv_p = V6 / reg / 'ncu_raw.csv'
        launches = parse_ncu_csv_v6(csv_p)
        # Aggregate per unique kernel name (mean across launches)
        by_kernel = defaultdict(list)
        for (kid, kname), info in launches.items():
            by_kernel[kname].append(info['metrics'])
        for kname, metrics_list in by_kernel.items():
            if not metrics_list:
                continue
            avg = {}
            for m_key in ('dram_throughput_pct', 'sm_throughput_pct',
                          'warps_active_pct', 'l1_throughput_pct',
                          'l2_throughput_pct', 'duration_ns'):
                vals = [m[m_key] for m in metrics_list if m_key in m]
                if vals:
                    avg[m_key] = round(sum(vals)/len(vals), 2)
                else:
                    avg[m_key] = 0
            rows.append({
                'model': 'lfm2.5-8b-a1b',
                'config': 'cookbook_baseline',
                'regime': reg,
                'batch_size': batch,
                'prompt_words': in_w,
                'output_tokens': out_t,
                'stage': stage,
                'kernel_short': short_name(kname),
                'kernel_full': kname[:80],
                'ncu_data_source': 'v6-jul-8-basic-sections',
                'sm_throughput_pct': avg.get('sm_throughput_pct', ''),
                'dram_throughput_pct': avg.get('dram_throughput_pct', ''),
                'tensor_pipe_active_pct': '',  # not in v6 basic
                'warps_active_pct': avg.get('warps_active_pct', ''),
                'l1_hit_pct': '',  # l1_throughput_pct not same as l1_hit_pct
                'l2_hit_pct': '',
                'stall_long_scoreboard_avg': '',
                'stall_math_pipe_throttle_avg': '',
                'verdict': '',
                'headroom_estimate_pct': '',
                'duration_ns_median': avg.get('duration_ns', ''),
            })
    return rows


def main():
    rows = load_june_qwen3_data() + load_v6_lfm_data()
    if not rows:
        print("No data found.")
        return
    # Sort
    rows.sort(key=lambda r: (r['model'], r['regime'], r['kernel_short']))

    # Write CSV
    out_csv = REPO / 'results/consolidated_v6_sglang_ncu.csv'
    # Union of all keys
    all_keys = list(rows[0].keys())
    for r in rows:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)
    with out_csv.open('w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=all_keys, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)
    print(f"Wrote {len(rows)} rows → {out_csv}")

    # xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
        wb = Workbook(); wb.remove(wb.active)

        # README
        ws = wb.create_sheet('README', 0)
        readme = [
            ('v6 sglang NCU kernel-level report', True),
            ('', False),
            ('Combines two data sources — both from REAL sglang kernels', True),
            ('', False),
            ('Data source 1: Qwen3-30B-A3B, June 9 experiment', True),
            ('  Method: sudo ncu --set full --launch-count 30 wrapping sglang.bench_one_batch', False),
            ('           with --profile-activities CUDA_PROFILER (sglang calls cudaProfilerStart', False),
            ('           inside so NCU only profiles the bench section)', False),
            ('  Coverage: 1 model (Qwen3-30B-A3B bf16) x 1 config (cookbook-equivalent) x 4 regimes', False),
            ('  Metrics: SM%, DRAM%, TensorCore%, Warps active (occupancy), L1/L2 hit%,', False),
            ('           stall counters, verdict + headroom estimate', False),
            ('', False),
            ('Data source 2: LFM2.5-8B-A1B, July 8 experiment', True),
            ('  Method: Same as v6 (bench_one_batch + NCU), but with --section SpeedOfLight', False),
            ('          + Occupancy + LaunchStats (subset for speed). Kernel-name filter to', False),
            ('          reduce number of profiled kernels.', False),
            ('  Coverage: 1 model (LFM2.5) x 1 config (cookbook) x 3 regimes', False),
            ('  Metrics: SM%, DRAM%, Occupancy%, L1/L2 throughput% (no hit%, no stalls)', False),
            ('', False),
            ('Both are from REAL sglang kernels (fused_moe_kernel, nvjet_tst_*, flashinfer::,', True),
            ('cutlass::flash::) — NOT from transformers. This is the difference from', False),
            ('v5b_ncu_report.xlsx which used transformers as a fallback.', False),
            ('', False),
            ('KEY findings visible in the data:', True),
            ('  * fused_moe_kernel is DRAM-bound (SM ~13-17%, DRAM 50-80%)', False),
            ('  * nvjet_tst GEMM kernels in decode: DRAM-bound (SM 8%, DRAM 40-50%)', False),
            ('  * BUT in PREFILL, nvjet_192x192 hits SM 95%, TC 96% — compute-bound!', False),
            ('    (This is why long_prefill is so different from decode.)', False),
            ('  * moe_sum_reduce hits DRAM 92% — bandwidth ceiling', False),
            ('  * Almost all kernels have low occupancy (5-15%) — Hopper SMs underused', False),
            ('    due to per-kernel register/shmem pressure', False),
            ('', False),
            ('Note on missing data:', True),
            ('  - No fp8 model NCU data (Qwen3-30B fp8 not in either source)', False),
            ('  - Only cookbook config profiled (no big_batch_cap128 NCU comparison)', False),
            ('  - LFM data only has SoL+Occupancy sections (no L1/L2 hit%, no stalls)', False),
        ]
        for i, (t, b) in enumerate(readme, 1):
            c = ws.cell(row=i, column=1, value=t)
            if b: c.font = Font(bold=True, size=11)
        ws.column_dimensions['A'].width = 105

        # Main sheet
        ws2 = wb.create_sheet('all_ncu_kernels')
        headers = all_keys
        hdr_fill = PatternFill('solid', fgColor='4A4A4A')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, '')
                ws2.cell(row=ri, column=ci, value=v)
        ws2.freeze_panes = 'I2'
        for ci, h in enumerate(headers, 1):
            col = get_column_letter(ci)
            width = min(max(len(h) + 2, 12),
                        40 if h == 'kernel_full' else 20)
            ws2.column_dimensions[col].width = width
            if 'pct' in h:
                for ri in range(2, len(rows) + 2):
                    cell = ws2[f'{col}{ri}']
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0"%"'
                # Color scale
                rng = f'{col}2:{col}{len(rows)+1}'
                rule = ColorScaleRule(
                    start_type='num', start_value=0, start_color='FFFFFF',
                    mid_type='num', mid_value=50, mid_color='FFEB84',
                    end_type='num', end_value=95, end_color='63BE7B',
                )
                ws2.conditional_formatting.add(rng, rule)

        # Hot kernels sheet (highest SM or DRAM)
        ws3 = wb.create_sheet('hot_kernels_by_regime')
        # Group by (model, regime), take top 5 by max(SM, DRAM)
        by_mr = defaultdict(list)
        for r in rows:
            by_mr[(r['model'], r['regime'])].append(r)
        cols3 = ['model', 'regime', 'kernel_short', 'sm_throughput_pct',
                 'dram_throughput_pct', 'tensor_pipe_active_pct',
                 'warps_active_pct', 'verdict', 'ncu_data_source']
        for ci, h in enumerate(cols3, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
        ri = 2
        for (mdl, reg), kernels in sorted(by_mr.items()):
            def score(k):
                try:
                    return max(float(k.get('sm_throughput_pct') or 0),
                                float(k.get('dram_throughput_pct') or 0))
                except:
                    return 0
            top = sorted(kernels, key=score, reverse=True)[:5]
            for k in top:
                for ci, h in enumerate(cols3, 1):
                    ws3.cell(row=ri, column=ci, value=k.get(h, ''))
                ri += 1
        ws3.freeze_panes = 'D2'
        for ci, h in enumerate(cols3, 1):
            col = get_column_letter(ci)
            ws3.column_dimensions[col].width = min(max(len(h) + 2, 14), 22)
            if 'pct' in h:
                for r_i in range(2, ri):
                    cell = ws3[f'{col}{r_i}']
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0"%"'

        out_xlsx = REPO / 'results/v6_sglang_ncu_report.xlsx'
        wb.save(out_xlsx)
        print(f'Wrote {out_xlsx}')
    except ImportError:
        print('openpyxl not installed')


if __name__ == '__main__':
    main()
