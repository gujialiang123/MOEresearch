#!/usr/bin/env python3
"""Aggregate v5b NCU kernel-level metrics into a clean spreadsheet.

For each (model, batch, kernel_group) we have an NCU CSV with rich per-kernel
metrics. We extract the KEY metrics and put them in one row per (kernel launch).
Then aggregate to one row per (model, batch, kernel_short_name) with median.

Key metrics extracted (from NCU --set full):
  - GPU Speed Of Light Throughput: DRAM Throughput%, SM Compute Throughput%,
    L1/TEX Cache Throughput%, L2 Cache Throughput%
  - Achieved Occupancy%
  - Duration (ns), Elapsed Cycles

Output:
  results/consolidated_v5b_ncu_kernels.csv
  results/v5b_ncu_report.xlsx
"""
from __future__ import annotations
import csv, json
from collections import defaultdict
from pathlib import Path
from statistics import median

REPO = Path(__file__).resolve().parent.parent
V5B = REPO / "results/2026-07-08_v5b_ncu"

MODELS = ["qwen3-0.6b", "qwen3-30b-a3b-bf16"]  # LFM excluded (transformers doesn't support lfm2_moe)
BATCHES = [1, 32, 128]
KERNEL_GROUPS = ["fused_moe", "flash_attn", "gemm"]

# Metrics to extract (Metric Name → column key)
KEY_METRICS = {
    "DRAM Throughput": "dram_throughput_pct",
    "Memory Throughput": "memory_throughput_pct",
    "Compute (SM) Throughput": "sm_compute_throughput_pct",
    "L1/TEX Cache Throughput": "l1_cache_throughput_pct",
    "L2 Cache Throughput": "l2_cache_throughput_pct",
    "Duration": "duration_ns",
    "Elapsed Cycles": "elapsed_cycles",
    "Achieved Occupancy": "achieved_occupancy_pct",
    "SM Active Cycles": "sm_active_cycles",
}


def parse_ncu_csv(csv_path: Path) -> list[dict]:
    """Parse NCU CSV output. Each row has one metric for one kernel launch.
    We pivot to one dict per kernel launch."""
    if not csv_path.exists():
        return []
    # NCU CSV has header + metric rows. Preamble may include non-CSV text
    # (like "==PROF== Connected..."). Skip until we hit header.
    text = csv_path.read_text()
    lines = text.splitlines()
    # Find "ID" header line
    header_idx = None
    for i, l in enumerate(lines):
        if l.startswith('"ID"'):
            header_idx = i
            break
    if header_idx is None:
        return []
    # Parse from that line
    csv_text = "\n".join(lines[header_idx:])
    reader = csv.DictReader(csv_text.splitlines())
    # Pivot: group by (ID, Kernel Name) → dict of metric_name: value
    launches = defaultdict(lambda: {"metrics": {}})
    for row in reader:
        kid = row.get("ID", "?")
        k = row.get("Kernel Name", "?")
        m = row.get("Metric Name", "")
        v = row.get("Metric Value", "")
        u = row.get("Metric Unit", "")
        launch_key = (kid, k)
        launches[launch_key]["kernel"] = k
        launches[launch_key]["id"] = kid
        if m in KEY_METRICS:
            try:
                launches[launch_key]["metrics"][KEY_METRICS[m]] = float(v)
            except (ValueError, TypeError):
                pass
    return list(launches.values())


def aggregate():
    all_rows = []
    for m in MODELS:
        for b in BATCHES:
            for kg in KERNEL_GROUPS:
                csv_p = V5B / m / f"batch_{b}" / kg / "ncu.csv"
                launches = parse_ncu_csv(csv_p)
                if not launches:
                    continue
                # Aggregate per unique kernel_name within this group
                by_kernel = defaultdict(list)
                for l in launches:
                    by_kernel[l["kernel"]].append(l["metrics"])
                for kn, metrics_list in by_kernel.items():
                    if not metrics_list:
                        continue
                    row = {
                        "model": m,
                        "batch_size": b,
                        "kernel_group": kg,
                        "kernel_name": kn[:80],
                        "n_launches_profiled": len(metrics_list),
                    }
                    for metric_key in KEY_METRICS.values():
                        vals = [m[metric_key] for m in metrics_list if metric_key in m]
                        if vals:
                            row[metric_key + "_median"] = round(median(vals), 3)
                        else:
                            row[metric_key + "_median"] = ""
                    all_rows.append(row)
    return all_rows


def main():
    rows = aggregate()
    # Sort by (model, batch, kernel_group)
    rows.sort(key=lambda r: (r["model"], r["batch_size"], r["kernel_group"], r["kernel_name"]))
    out_csv = REPO / "results/consolidated_v5b_ncu_kernels.csv"
    if rows:
        with out_csv.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
    print(f"Wrote {len(rows)} rows → {out_csv}")

    # xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
        wb = Workbook()
        wb.remove(wb.active)

        # README
        ws = wb.create_sheet("README", 0)
        readme = [
            ("v5b NCU kernel-level hardware counters", True),
            ("", False),
            ("Method:", True),
            ("  1. For each (model, batch_size, kernel_group), spawn a Python script that:", False),
            ("     - loads the model in bf16 via transformers", False),
            ("     - does 3 warmup forward passes (build KV cache)", False),
            ("     - runs 5 decode steps at the target batch size", False),
            ("  2. Wrap the script in Nsight Compute (NCU) with:", False),
            ("     -k regex:<kernel_group_pattern> -c 5 --set full", False),
            ("     This makes NCU profile only 5 launches of matching kernels,", False),
            ("     collecting ~100 hardware metrics per launch.", False),
            ("  3. NCU takes ~30-70s per (model, batch, kernel) combo because", False),
            ("     it replays each kernel N times to gather different metric sections.", False),
            ("", False),
            ("Coverage:", True),
            ("  2 models (Qwen3-0.6B, Qwen3-30B-A3B bf16)", False),
            ("  × 3 batch sizes (1, 32, 128)", False),
            ("  × 3 kernel groups (fused_moe, flash_attn, gemm)", False),
            ("  = 18 combos, actual launches profiled = 5 per matched kernel", False),
            ("", False),
            ("NOTE: LFM2.5-8B-A1B NOT included here.", True),
            ("  Its 'lfm2_moe' arch is not supported by transformers 4.57 (only by sglang).", False),
            ("  For LFM2.5 kernel data see results/consolidated_v5_kernel_breakdown.csv", False),
            ("  which uses kineto (from sglang server directly).", False),
            ("", False),
            ("Key metrics (all are peak-fraction %):", True),
            ("  dram_throughput_pct    = HBM read+write bandwidth / peak_HBM_BW", False),
            ("                           MOST IMPORTANT for decode (memory-bound)", False),
            ("  sm_compute_throughput_pct = tensor core + FP32/INT32 pipes utilized", False),
            ("                               / peak. Real MFU signal.", False),
            ("  l1_cache_throughput_pct, l2_cache_throughput_pct = cache bandwidth", False),
            ("                                                     usage", False),
            ("  achieved_occupancy_pct = actively resident warps / theoretical max", False),
            ("  duration_ns            = median kernel duration for this batch size", False),
            ("", False),
            ("Sanity check patterns to look for:", True),
            ("  1. dram_throughput_pct should INCREASE with batch size for GEMM", False),
            ("     (larger reads → better bandwidth utilization)", False),
            ("  2. sm_compute_throughput_pct should INCREASE with batch size", False),
            ("     (arithmetic intensity grows → tensor core gets fuller)", False),
            ("  3. duration_ns should scale sub-linearly with batch (batching amortizes)", False),
            ("  4. flash_attn dram% is typically much lower than gemm (compute-bound", False),
            ("     due to O(N²) attention math)", False),
        ]
        hdr_fill = PatternFill('solid', fgColor='4A4A4A')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for i, (t, bold) in enumerate(readme, 1):
            c = ws.cell(row=i, column=1, value=t)
            if bold:
                c.font = Font(bold=True, size=11)
        ws.column_dimensions['A'].width = 110

        # Sheet: full breakdown
        if rows:
            ws2 = wb.create_sheet("ncu_kernels")
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = Alignment(horizontal='center', wrap_text=True)
            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    ws2.cell(row=ri, column=ci, value=row.get(h))
            ws2.freeze_panes = 'E2'
            for ci, h in enumerate(headers, 1):
                col = get_column_letter(ci)
                width = min(max(len(h) + 2, 12),
                            40 if h == 'kernel_name' else 22)
                ws2.column_dimensions[col].width = width
                # Format pct cols
                if 'pct' in h:
                    for ri in range(2, len(rows) + 2):
                        cell = ws2[f"{col}{ri}"]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00"%"'
                    # Color scale
                    rng = f"{col}2:{col}{len(rows)+1}"
                    rule = ColorScaleRule(
                        start_type='num', start_value=0, start_color='FFFFFF',
                        mid_type='num', mid_value=50, mid_color='FFEB84',
                        end_type='num', end_value=90, end_color='63BE7B',
                    )
                    ws2.conditional_formatting.add(rng, rule)

        out_xlsx = REPO / "results/v5b_ncu_report.xlsx"
        wb.save(out_xlsx)
        print(f"Wrote {out_xlsx}")
    except ImportError:
        print("openpyxl not installed; skipping xlsx")


if __name__ == "__main__":
    main()
