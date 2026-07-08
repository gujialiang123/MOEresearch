#!/usr/bin/env python3
"""Aggregate v5 kineto kernel-level data into a shareable spreadsheet.

Output:
  results/consolidated_v5_kernel_breakdown.csv
    rows: (model, config, regime, kernel_name)
    cols: total_time_ms, count, mean_us, % of GPU time

  results/v5_kernel_report.xlsx
    Sheet 1: kernel_breakdown (full data)
    Sheet 2: top_kernel_summary (top-3 per combo, wide)
    Sheet 3: README
"""
from __future__ import annotations
import csv, json
from collections import defaultdict
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
V5 = REPO / "results/2026-07-08_v5_ncu"

MODELS = ['lfm2.5-8b-a1b', 'qwen3-30b-a3b-bf16']
CONFIGS = ['cookbook_baseline', 'big_batch_cap128']
REGIMES = ['R_decode_c1_out2k', 'R_conc_ref', 'R_decode_c128_out256']


def short_name(kname: str) -> str:
    """Turn a long CUDA kernel name into a human-readable short label."""
    kname_lower = kname.lower()
    if 'fused_moe_kernel' in kname_lower:
        return 'fused_moe_kernel'
    if 'flash::' in kname_lower or ('cutlass' in kname_lower and 'flash' in kname_lower):
        return 'flash_attention'
    if 'flashinfer::norm::fusedaddrmsnorm' in kname_lower or 'FusedAddRMSNorm' in kname:
        return 'flashinfer_FusedAddRMSNorm'
    if 'flashinfer::norm::rmsnorm' in kname_lower:
        return 'flashinfer_RMSNorm'
    if 'flashinfer::norm' in kname_lower:
        return 'flashinfer_norm'
    if 'flashinfer::gemm' in kname_lower:
        return 'flashinfer_gemm'
    if 'flashinfer::sampling' in kname_lower:
        return 'flashinfer_sampling'
    if 'nvjet_tst' in kname_lower:
        # Extract the shape signature
        parts = kname.split('_')
        if len(parts) >= 4:
            return f'nvjet_gemm_{parts[2]}_{parts[3]}'
        return 'nvjet_gemm'
    if 'elementwise' in kname_lower:
        if 'BinaryFunctor' in kname:
            if 'MulFunctor' in kname: return 'elementwise_mul'
            if 'AddFunctor' in kname: return 'elementwise_add'
            if 'SubFunctor' in kname: return 'elementwise_sub'
            return 'elementwise_binary'
        return 'elementwise'
    if 'rope' in kname_lower or 'rotary' in kname_lower:
        return 'rope'
    if 'softmax' in kname_lower:
        return 'softmax'
    if 'reduce' in kname_lower:
        return 'reduce'
    if 'copy' in kname_lower:
        return 'memcpy'
    # Default: first 40 chars
    return kname[:40]


def main():
    csv_rows = []
    summary_rows = []
    for m in MODELS:
        for c in CONFIGS:
            for r in REGIMES:
                p = V5 / m / c / r / 'kineto_top_kernels.json'
                if not p.exists():
                    continue
                d = json.load(open(p))
                # Get bench result
                bench = d.get('bench_result', {}).get('regimes', {}).get(r, {})
                tps = bench.get('tokens_per_s', {}).get('mean', 0)
                rps = bench.get('req_per_s', {}).get('mean', 0)
                # Compute total kernel time and per-kernel share
                total_us = sum(k['total_us'] for k in d.get('top_kernels', []))
                for i, k in enumerate(d.get('top_kernels', [])):
                    sn = short_name(k['name'])
                    share = 100 * k['total_us'] / total_us if total_us > 0 else 0
                    csv_rows.append({
                        'model': m,
                        'config': c,
                        'regime': r,
                        'bench_tokens_per_s': round(tps, 1),
                        'bench_req_per_s': round(rps, 3),
                        'kernel_rank': i + 1,
                        'kernel_short': sn,
                        'kernel_full_name': k['name'][:100],
                        'count': k['count'],
                        'total_ms': round(k['total_us'] / 1000, 2),
                        'mean_us': round(k['mean_us'], 2),
                        'share_of_top5_pct': round(share, 1),
                    })
                # Wide summary — top 3
                sum_row = {
                    'model': m, 'config': c, 'regime': r,
                    'bench_tokens_per_s': round(tps, 1),
                    'bench_req_per_s': round(rps, 3),
                }
                for i in range(3):
                    if i < len(d.get('top_kernels', [])):
                        k = d['top_kernels'][i]
                        sn = short_name(k['name'])
                        sum_row[f'k{i+1}_name'] = sn
                        sum_row[f'k{i+1}_total_ms'] = round(k['total_us'] / 1000, 2)
                        sum_row[f'k{i+1}_count'] = k['count']
                        sum_row[f'k{i+1}_mean_us'] = round(k['mean_us'], 2)
                summary_rows.append(sum_row)

    # Write CSV
    out_csv = REPO / 'results/consolidated_v5_kernel_breakdown.csv'
    if csv_rows:
        with out_csv.open('w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(csv_rows[0].keys()))
            w.writeheader()
            w.writerows(csv_rows)
    print(f"Wrote {len(csv_rows)} rows → {out_csv}")

    out_summary = REPO / 'results/consolidated_v5_kernel_summary.csv'
    if summary_rows:
        with out_summary.open('w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
            w.writeheader()
            w.writerows(summary_rows)
    print(f"Wrote {len(summary_rows)} rows → {out_summary}")

    # xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        wb.remove(wb.active)

        # README
        ws = wb.create_sheet('README', 0)
        readme = [
            ('v5 kernel-level profiling — kineto traces from sglang', True),
            ('', False),
            ('Method:', True),
            ('  1. Launch sglang server with SGLANG_TORCH_PROFILER_DIR set', False),
            ('  2. Send 32-request warmup burst to steady-state the server', False),
            ('  3. Call /start_profile → kineto starts capturing', False),
            ('  4. Run the target regime bench (3 runs)', False),
            ('  5. Call /stop_profile → trace written to disk', False),
            ('  6. Parse trace, aggregate per-kernel total GPU time', False),
            ('', False),
            ('12 combos: 2 models × 2 configs × 3 regimes', True),
            ('  models: LFM2.5-8B-A1B (bf16), Qwen3-30B-A3B (bf16)', False),
            ('  configs: cookbook_baseline (cap=32, chunk=-1), big_batch_cap128 (cap=128, chunk=2048)', False),
            ('  regimes: R_decode_c1_out2k (batch=1), R_conc_ref (batch=32), R_decode_c128_out256 (batch=128)', False),
            ('', False),
            ('Sheets:', True),
            ('  kernel_breakdown: one row per (combo, kernel_rank 1-5)', False),
            ('  top_kernel_summary: one row per combo, top 3 kernels side-by-side', False),
            ('', False),
            ('Key columns:', True),
            ('  bench_tokens_per_s: MEASURED sglang throughput', False),
            ('  count: number of times this kernel was launched during the bench', False),
            ('  total_ms: total GPU time in this kernel during the bench', False),
            ('  mean_us: average kernel duration per launch', False),
            ('  share_of_top5_pct: % of top-5 kernels total GPU time', False),
            ('', False),
            ('Sanity checks worth doing:', True),
            ('  1. fused_moe_kernel dominates in all MoE combos (60-80% of GPU time)', False),
            ('  2. Same model, same regime: cookbook vs big_batch has near-identical kernels', False),
            ('     if regime uses batch ≤ 32 (both fit in either config)', False),
            ('  3. At batch=128 (R_decode_c128), big_batch config uses larger nvjet GEMM shapes', False),
            ('  4. Per-launch times scale with batch (larger batch = larger GEMM = longer per-call)', False),
        ]
        for i, (t, bold) in enumerate(readme, 1):
            c = ws.cell(row=i, column=1, value=t)
            if bold:
                c.font = Font(bold=True, size=11)
        ws.column_dimensions['A'].width = 110

        # Sheet: full breakdown
        ws2 = wb.create_sheet('kernel_breakdown')
        headers = list(csv_rows[0].keys())
        hdr_fill = PatternFill('solid', fgColor='4A4A4A')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(csv_rows, 2):
            for ci, h in enumerate(headers, 1):
                ws2.cell(row=ri, column=ci, value=row.get(h))
        ws2.freeze_panes = 'D2'
        for ci, h in enumerate(headers, 1):
            col = get_column_letter(ci)
            width = min(max(len(h) + 2, 14), 40 if h == 'kernel_full_name' else 20)
            ws2.column_dimensions[col].width = width

        # Sheet: top 3 kernel summary (wide)
        ws3 = wb.create_sheet('top_kernel_summary')
        headers3 = list(summary_rows[0].keys())
        for ci, h in enumerate(headers3, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(summary_rows, 2):
            for ci, h in enumerate(headers3, 1):
                ws3.cell(row=ri, column=ci, value=row.get(h))
        ws3.freeze_panes = 'D2'
        for ci, h in enumerate(headers3, 1):
            col = get_column_letter(ci)
            ws3.column_dimensions[col].width = min(max(len(h) + 2, 12), 28)

        xlsx = REPO / 'results/v5_kernel_report.xlsx'
        wb.save(xlsx)
        print(f"Wrote {xlsx}")
    except ImportError:
        print("openpyxl not installed; skipping xlsx")


if __name__ == '__main__':
    main()
